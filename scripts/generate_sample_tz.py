"""Generate test_data/sample_tz.xlsx with sample customer requirements.

Run from repo root:

    python3 scripts/generate_sample_tz.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "ID",
    "Требование заказчика",
    "Ожидаемый статус",
    "Комментарий эксперта (эталон)",
]

ROWS = [
    (1, "Интеграция с Битрикс24 для синхронизации контактов", "Да", "Поддерживается коннектором v2.1"),
    (2, "Анализ эмоций абонентов в реальном времени", "Частично", "Только постобработка, не реалтайм"),
    (3, "Прямая интеграция с SAP через RFC", "НД", "Нет данных в публичной документации"),
    (4, "Запись всех разговоров в облаке", "Да", "Стандартная функция тарифа Профи"),
    (5, "API для внешней CRM с методом POST /call/start", "Да", "Описано в Public API Guide v2.0"),
]


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ROWS:
        ws.append(row)

    widths = [6, 60, 18, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    return wb


def main() -> Path:
    out_path = Path(__file__).resolve().parents[1] / "test_data" / "sample_tz.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
